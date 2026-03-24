import streamlit as st
import pandas as pd
import io
import datetime as dt
from datetime import time
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import plotly.express as px

# ─────────────────────────────────────────────
# PAGE CONFIG
# ─────────────────────────────────────────────
st.set_page_config(
    page_title="Attendance Report System",
    page_icon="📋",
    layout="wide",
    initial_sidebar_state="expanded"
)

st.markdown("""
<style>
    .main-header {
        background: linear-gradient(135deg, #1A3C6E, #2D6099);
        color: white; padding: 20px 30px; border-radius: 10px;
        margin-bottom: 20px; text-align: center;
    }
    .issue-critical {
        background: #FFEBEE; border-left: 5px solid #B71C1C;
        padding: 10px 16px; border-radius: 6px; margin: 6px 0;
    }
    .issue-box {
        background: #FFF3E0; border-left: 5px solid #E65100;
        padding: 10px 16px; border-radius: 6px; margin: 6px 0;
    }
    .issue-info {
        background: #E3F2FD; border-left: 5px solid #1565C0;
        padding: 10px 16px; border-radius: 6px; margin: 6px 0;
    }
    div[data-testid="stDataFrame"] { border-radius: 8px; overflow: hidden; }
    .stTabs [data-baseweb="tab"] { font-size: 13px; font-weight: 600; }
</style>
""", unsafe_allow_html=True)

# ─────────────────────────────────────────────
# CONSTANTS & HELPERS
# ─────────────────────────────────────────────
SUNDAYS   = {1,8,15,22}
SATURDAYS = {7,14,21,28}
DAY_ABBR  = {d: dt.date(2026,2,d).strftime('%a') for d in range(1,29)}
GRACE_F   = 5.0/1440.0
OT_F      = 29.0/1440.0

def parse_time_str(t):
    if pd.isna(t): return None
    s = str(t).strip()
    if ':' in s:
        try:
            p = s.split(':')
            return time(int(p[0]), int(p[1]))
        except: return None
    return None

def parse_time_cell(val):
    if pd.isna(val) or val == '' or val == 0: return None
    s = str(val).strip()
    if s in ('0','0.0','nan',''): return None
    if ':' in s and len(s) <= 5:
        try:
            p = s.split(':')
            h, m = int(p[0]), int(p[1])
            if 0 <= h <= 23 and 0 <= m <= 59:
                return time(h, m)
        except: pass
    if hasattr(val, 'hour'):
        return time(val.hour, val.minute)
    return None

def time_val(t):
    if t is None: return None
    return (t.hour*60 + t.minute) / 1440.0

def t2m(t):
    if t is None: return 0
    return t.hour*60 + t.minute

def m2hm(m):
    if not m or m <= 0: return "0:00"
    return f"{int(m//60)}:{int(m%60):02d}"

# ─────────────────────────────────────────────
# CORE CALCULATION
# ─────────────────────────────────────────────
def calc_day(d, in_t, out_t, si_t, so_t, is_open=False, std_hours=9.5):
    is_sun = d in SUNDAYS
    r = dict(status='', work=0, ot=0, late=0, early=0)
    if in_t is None and out_t is None:
        r['status'] = 'WO' if is_sun else 'A'; return r
    if in_t is None or out_t is None:
        r['status'] = 'MISS'; return r
    in_m  = t2m(in_t)
    out_m = t2m(out_t)
    if is_sun:
        cap = 16*60
        if is_open:
            r['work'] = max(0, min(out_m, cap) - in_m)
        else:
            si_m = t2m(si_t) if si_t else 0
            eff_in_m = si_m if in_m <= si_m + 5 else in_m
            eff_in_m = max(si_m, eff_in_m)
            r['work'] = max(0, min(out_m, cap) - eff_in_m)
        r['status'] = 'WOP'; return r
    if is_open:
        r['work']  = max(0, out_m - in_m)
        std_mins   = int(std_hours * 60)
        extra      = r['work'] - std_mins
        r['ot']    = extra if extra > 29 else 0
        r['status'] = 'P'
    else:
        si_m = t2m(si_t) if si_t else 0
        so_m = t2m(so_t) if so_t else 0
        eff_in_m = si_m if in_m <= si_m + 5 else in_m
        eff_in_m = max(si_m, eff_in_m)
        r['work']  = max(0, min(out_m, so_m) - eff_in_m)
        r['late']  = max(0, eff_in_m - si_m) if eff_in_m > si_m else 0
        r['early'] = max(0, so_m - out_m) if out_m < so_m else 0
        extra = out_m - so_m
        r['ot']    = extra if extra > 29 else 0
        r['status'] = 'P'
    return r

# ─────────────────────────────────────────────
# LOAD MASTER
# ─────────────────────────────────────────────
def load_master(file):
    df = pd.read_excel(file, sheet_name=0, header=None)
    df = df.dropna(axis=1, how='all')
    expected_cols = ['EmpCode','EmpName','Company','Department','Hour','Timing','ShiftIn','ShiftOut','Shift']
    actual_cols = len(df.columns)
    if actual_cols >= len(expected_cols):
        df = df.iloc[:, :len(expected_cols)]
        df.columns = expected_cols
    else:
        df.columns = expected_cols[:actual_cols]
    df = df.iloc[1:].reset_index(drop=True)
    df = df.dropna(subset=['EmpCode'], how='all').reset_index(drop=True)
    df['ShiftInTime']  = df['ShiftIn'].apply(parse_time_str)  if 'ShiftIn'  in df.columns else None
    df['ShiftOutTime'] = df['ShiftOut'].apply(parse_time_str) if 'ShiftOut' in df.columns else None
    def safe_int(x):
        try: return int(float(str(x)))
        except: return None
    df['EmpCode'] = df['EmpCode'].apply(safe_int)
    df = df[df['EmpCode'].notna()].reset_index(drop=True)
    def is_open(row):
        shift_val = str(row.get('Shift','')).strip().lower()
        if 'open' in shift_val: return True
        if pd.isna(row.get('ShiftInTime')) or pd.isna(row.get('ShiftOutTime')): return True
        return False
    df['IsOpen'] = df.apply(is_open, axis=1)
    def get_std_hours(row):
        try: return float(str(row['Hour']).strip())
        except: return 9.5
    df['StdHours'] = df.apply(get_std_hours, axis=1)
    return df

# ─────────────────────────────────────────────
# LOAD ATTENDANCE
# ─────────────────────────────────────────────
def load_attendance(file):
    att = {}
    try: xl = pd.ExcelFile(file)
    except: return att
    for sheet_name in xl.sheet_names:
        try: df = pd.read_excel(file, sheet_name=sheet_name, header=None)
        except: continue
        global_day_col_map = None
        for ri in range(len(df)):
            if str(df.iloc[ri, 0]).strip() == 'Days':
                global_day_col_map = {}
                for ci in range(len(df.iloc[ri])):
                    cv = str(df.iloc[ri, ci]).strip().split()
                    if cv and cv[0].isdigit():
                        global_day_col_map[ci] = int(cv[0])
                break
        i = 0
        while i < len(df):
            if str(df.iloc[i, 0]).strip() == 'Emp. Code:':
                emp_code = None
                for j in range(1, len(df.iloc[i])):
                    v = df.iloc[i, j]
                    if not pd.isna(v) and str(v).strip() not in ['','nan','Emp. Name:']:
                        try: emp_code = int(float(str(v))); break
                        except: pass
                if emp_code is None or i+3 >= len(df): i += 1; continue
                next_label = str(df.iloc[i+1, 0]).strip()
                if next_label == 'Status':
                    day_col_map = global_day_col_map
                    in_row = df.iloc[i+2]; out_row = df.iloc[i+3]
                else:
                    day_col_map = {}
                    for ci in range(len(df.iloc[i+1])):
                        cv = str(df.iloc[i+1, ci]).strip().split()
                        if cv and cv[0].isdigit():
                            day_col_map[ci] = int(cv[0])
                    in_row = df.iloc[i+2]; out_row = df.iloc[i+3]
                if emp_code not in att: att[emp_code] = {}
                if day_col_map:
                    for ci, dn in day_col_map.items():
                        if ci < len(in_row) and ci < len(out_row):
                            att[emp_code][dn] = {
                                'in':  parse_time_cell(in_row.iloc[ci]),
                                'out': parse_time_cell(out_row.iloc[ci])
                            }
                i += 4
            else:
                i += 1
    return att

# ─────────────────────────────────────────────
# DATA QUALITY CHECKS
# ─────────────────────────────────────────────
def run_data_checks(emp_df, attendance):
    issues = []
    master_codes = set(emp_df['EmpCode'].tolist())
    att_codes    = set(attendance.keys())

    # 1. In attendance but NOT in master — salary critical
    for c in sorted(att_codes - master_codes):
        issues.append({'severity':'🔴 CRITICAL','category':'Missing in Master',
                       'emp_code':c,'emp_name':'—',
                       'detail':f"Employee {c} has attendance data but is NOT in Employee Master. "
                                f"Work hours CANNOT be calculated. Add to master immediately."})

    # 2. In master but no attendance
    for c in sorted(master_codes - att_codes):
        row = emp_df[emp_df['EmpCode']==c]
        name = row.iloc[0]['EmpName'] if not row.empty else '?'
        issues.append({'severity':'🟡 WARNING','category':'No Attendance Data',
                       'emp_code':c,'emp_name':name,
                       'detail':f"{name} ({c}) is in master but has NO attendance records. Will show Absent all days."})

    # 3. Shift config errors
    for _, emp in emp_df.iterrows():
        code=emp['EmpCode']; name=str(emp['EmpName'])
        is_open=bool(emp.get('IsOpen',False))
        si=emp.get('ShiftInTime'); so=emp.get('ShiftOutTime')
        std=emp.get('StdHours',9.5)
        if not is_open and (pd.isna(si) or pd.isna(so)):
            issues.append({'severity':'🔴 CRITICAL','category':'Shift Config Error',
                           'emp_code':code,'emp_name':name,
                           'detail':f"{name} ({code}) has no Shift In/Out and is NOT Open Shift. Duration CANNOT be calculated."})
        if not is_open and si and so and t2m(si) >= t2m(so):
            issues.append({'severity':'🔴 CRITICAL','category':'Shift Time Error',
                           'emp_code':code,'emp_name':name,
                           'detail':f"{name} ({code}) Shift In={si.strftime('%H:%M')} >= Out={so.strftime('%H:%M')}. Wrong calculations will result."})
        if is_open and (float(std) <= 0 or float(std) > 24):
            issues.append({'severity':'🟡 WARNING','category':'Std Hours Issue',
                           'emp_code':code,'emp_name':name,
                           'detail':f"{name} ({code}) Open Shift but StdHours={std} is invalid. OT will be wrong."})

    # 4. MISS punches
    for _, emp in emp_df.iterrows():
        code=emp['EmpCode']; name=str(emp['EmpName'])
        emp_att=attendance.get(code,{})
        miss_days=[d for d in range(1,29) if
                   (emp_att.get(d,{}).get('in') is None) != (emp_att.get(d,{}).get('out') is None)]
        if miss_days:
            issues.append({'severity':'🟠 MISS','category':'Missing Punch',
                           'emp_code':code,'emp_name':name,
                           'detail':f"{name} ({code}) missing punch on: {', '.join(f'Feb {d}' for d in miss_days)}. Hours not counted."})

    # 5. Suspicious punch times
    for _, emp in emp_df.iterrows():
        code=emp['EmpCode']; name=str(emp['EmpName'])
        emp_att=attendance.get(code,{})
        suspect=[]
        for d in range(1,29):
            if d in SUNDAYS: continue
            dd=emp_att.get(d,{})
            in_t=dd.get('in'); out_t=dd.get('out')
            if in_t and t2m(in_t)>14*60: suspect.append(f"Feb {d} In={in_t.strftime('%H:%M')}")
            if out_t and t2m(out_t)<10*60: suspect.append(f"Feb {d} Out={out_t.strftime('%H:%M')}")
        if suspect:
            issues.append({'severity':'🔵 INFO','category':'Suspicious Punch',
                           'emp_code':code,'emp_name':name,
                           'detail':f"{name} ({code}) unusual times: {', '.join(suspect[:5])}"
                                    +(' ...' if len(suspect)>5 else '')})
    return issues

# ─────────────────────────────────────────────
# GATE PASS HELPER
# ─────────────────────────────────────────────
def get_gp_deduction(gate_passes, code, day):
    return sum(gp['duration_mins'] for gp in gate_passes if gp['code']==code and gp['day']==day)

# ─────────────────────────────────────────────
# COMPUTE SUMMARY
# ─────────────────────────────────────────────
def compute_summary(emp_df, attendance, gate_passes=None):
    if gate_passes is None: gate_passes=[]
    rows=[]
    for _,emp in emp_df.iterrows():
        try: code=int(float(str(emp['EmpCode'])))
        except: code=str(emp['EmpCode'])
        si_t=emp['ShiftInTime']; so_t=emp['ShiftOutTime']
        is_open=bool(emp.get('IsOpen',False)); std_hours=float(emp.get('StdHours',9.5))
        emp_att=attendance.get(code,{})
        T=dict(P=0,WOP=0,WO=0,A=0,MISS=0,work=0,ot=0,late=0,early=0,late_days=0,early_days=0,gp=0,gp_days=0)
        for d in range(1,29):
            dd=emp_att.get(d,{})
            r=calc_day(d,dd.get('in'),dd.get('out'),si_t,so_t,is_open,std_hours)
            gp_mins=get_gp_deduction(gate_passes,code,d)
            if r['status'] in T: T[r['status']]+=1
            T['work']+=max(0,r['work']-gp_mins); T['ot']+=r['ot']
            T['late']+=r['late']; T['early']+=r['early']
            T['gp']+=gp_mins
            if gp_mins>0: T['gp_days']+=1
            if r['late']>0: T['late_days']+=1
            if r['early']>0: T['early_days']+=1
        avg=T['work']/max(T['P']+T['WOP'],1)
        net=max(0,T['work']+T['ot']-T['late'])
        shift_lbl=('Open Shift' if is_open
                   else f"{si_t.strftime('%H:%M') if si_t else '?'}–{so_t.strftime('%H:%M') if so_t else '?'}")
        rows.append({'Code':code,'Name':str(emp['EmpName']),'Company':str(emp['Company']),
                     'Department':str(emp['Department']),'Shift':shift_lbl,
                     'Type':'🔓 Open' if is_open else '🔒 Fixed',
                     'Present':T['P'],'WOP':T['WOP'],'W/Off':T['WO'],'Absent':T['A'],'MISS':T['MISS'],
                     'Work Hrs':m2hm(T['work']),'OT Hrs':m2hm(T['ot']),
                     'Late Hrs':m2hm(T['late']),'Late Days':T['late_days'],
                     'Early Hrs':m2hm(T['early']),'Early Days':T['early_days'],
                     'Gate Pass Hrs':m2hm(T['gp']),'GP Days':T['gp_days'],
                     'Dur+OT':m2hm(T['work']+T['ot']),'Net Work Hrs':m2hm(net),
                     'Avg Hrs/Day':m2hm(int(avg)),
                     '_work_mins':T['work'],'_ot_mins':T['ot'],'_present_total':T['P']+T['WOP']})
    return pd.DataFrame(rows)

# ─────────────────────────────────────────────
# BUILD ATTENDANCE EXCEL
# ─────────────────────────────────────────────
def build_excel(emp_df, attendance, gate_passes=None):
    if gate_passes is None: gate_passes=[]
    thin=Side(style='thin'); med=Side(style='medium')
    def tb(): return Border(left=thin,right=thin,top=thin,bottom=thin)
    def mb(): return Border(left=med,right=med,top=med,bottom=med)
    def fl(h): return PatternFill('solid',start_color=h)
    def fn(bold=False,size=9,color='000000'): return Font(name='Arial',bold=bold,size=size,color=color)
    def al(h='center',v='center',wrap=True): return Alignment(horizontal=h,vertical=v,wrap_text=wrap)

    SHIFTIN_COL=2; SHIFTOUT_COL=3; DAY1_COL=4; DAY_END=31; SUM1_COL=32
    SUM_LABELS=['Present','WOP','W/Off','Absent','MISS','Total Work\nDuration','Total OT\nHrs',
                'Late By\nHrs','Late By\nDays','Early By\nHrs','Early By\nDays',
                'Total Dur\n(+OT)','Avg Work\nHrs','Net Work\nHrs\n(Work+OT\n−Late)','Total\nDur+OT\n(All Days)']
    LAST_COL=SUM1_COL+len(SUM_LABELS)-1

    wb=Workbook(); wb.remove(wb.active)
    for company in emp_df['Company'].unique():
        c_emps=emp_df[emp_df['Company']==company].copy()
        if c_emps.empty: continue
        ws=wb.create_sheet(title=str(company)[:31])
        for r,txt,sz in [(1,"Monthly Status Report (Detailed Work Duration)",14),(2,"Feb 01 2026  To  Feb 28 2026",10)]:
            ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=LAST_COL)
            c=ws.cell(r,1,txt); c.font=fn(True,sz,'FFFFFF'); c.fill=fl('1A3C6E'); c.alignment=al()
            ws.row_dimensions[r].height=22 if r==1 else 16
        ws.merge_cells(start_row=3,start_column=1,end_row=3,end_column=LAST_COL)
        ws.cell(3,1).fill=fl('1A3C6E'); ws.row_dimensions[3].height=4
        ws.merge_cells(start_row=4,start_column=1,end_row=4,end_column=8)
        c=ws.cell(4,1,f"Company:  {company}"); c.font=fn(True,10); c.alignment=al('left')
        ws.row_dimensions[4].height=16
        ws.row_dimensions[5].height=32
        for col,txt2,fg,bg in [(1,'Days','000000','EEF4FA'),(2,'Sh.In','1565C0','DDEEFF'),(3,'Sh.Out','1565C0','DDEEFF')]:
            c=ws.cell(5,col,txt2); c.font=fn(True,8,fg); c.fill=fl(bg); c.alignment=al(); c.border=tb()
        for d in range(1,29):
            col=DAY1_COL+d-1
            c=ws.cell(5,col,f"{d}\n{DAY_ABBR[d][0]}")
            if d in SUNDAYS: c.fill=fl('FFA726'); c.font=fn(True,8,'7B1FA2')
            elif d in SATURDAYS: c.fill=fl('A5D6A7'); c.font=fn(True,8,'1B5E20')
            else: c.fill=fl('D0E4F5'); c.font=fn(True,8)
            c.alignment=al(wrap=True); c.border=tb()
        for si2,sh in enumerate(SUM_LABELS):
            col=SUM1_COL+si2; c=ws.cell(5,col,sh)
            c.font=fn(True,8,'FFFFFF')
            c.fill=fl('4A235A') if si2==14 else fl('1F4E79')
            c.alignment=al(wrap=True); c.border=tb()

        current_row=6
        for dept in c_emps['Department'].unique():
            d_emps=c_emps[c_emps['Department']==dept]
            ws.merge_cells(start_row=current_row,start_column=1,end_row=current_row,end_column=LAST_COL)
            c=ws.cell(current_row,1,f"  Department:  {dept}")
            c.font=fn(True,10,'FFFFFF'); c.fill=fl('2D5F8A'); c.alignment=al('left')
            ws.row_dimensions[current_row].height=16; current_row+=1

            for _,emp in d_emps.iterrows():
                try: emp_code=int(float(str(emp['EmpCode'])))
                except: emp_code=str(emp['EmpCode'])
                emp_name=str(emp['EmpName']); si_t=emp['ShiftInTime']; so_t=emp['ShiftOutTime']
                is_open=bool(emp.get('IsOpen',False)); std_hours=float(emp.get('StdHours',9.5))
                std_mins_frac=(std_hours*60)/1440.0
                siv=time_val(si_t); sov=time_val(so_t)
                if is_open: shift_str=f"Open Shift  (Std: {std_hours}h)"
                else: shift_str=(f"{si_t.strftime('%H:%M')}–{so_t.strftime('%H:%M')}"
                                 if si_t and so_t else str(emp.get('Shift','Open')))
                emp_att=attendance.get(emp_code,{})
                HR=current_row; RS=HR+1; RI=HR+2; RO=HR+3; RD=HR+4
                RL=HR+5; RE=HR+6; ROT=HR+7; RDP=HR+8; RSP=HR+9
                for rr,h in [(HR,18),(RS,14),(RI,14),(RO,14),(RD,14),(RL,14),(RE,14),(ROT,14),(RDP,14),(RSP,3)]:
                    ws.row_dimensions[rr].height=h
                c=ws.cell(HR,1,f"  {emp_code} : {emp_name}")
                c.font=fn(True,9); c.fill=fl('E8F0E8' if is_open else 'E8F4FD'); c.alignment=al('left'); c.border=tb()
                c=ws.cell(HR,SHIFTIN_COL)
                if is_open: c.value=None; c.font=fn(False,8,'2E7D32'); c.fill=fl('E8F0E8')
                else: c.value=siv; c.number_format='h:mm'; c.font=fn(False,8,'1565C0'); c.fill=fl('DDEEFF')
                c.alignment=al(); c.border=tb()
                c=ws.cell(HR,SHIFTOUT_COL)
                if is_open: c.value=std_mins_frac; c.number_format='[h]:mm'; c.font=fn(False,8,'2E7D32'); c.fill=fl('E8F0E8')
                else: c.value=sov; c.number_format='h:mm'; c.font=fn(False,8,'1565C0'); c.fill=fl('DDEEFF')
                c.alignment=al(); c.border=tb()
                ws.merge_cells(start_row=HR,start_column=4,end_row=HR,end_column=LAST_COL)
                hint=("Open Shift — Duration=Out−In | OT if Duration>Std Hrs | No Late/Early" if is_open
                      else "Fill MISS punch in InTime/OutTime rows ↓  |  All formulas auto-recalculate")
                c=ws.cell(HR,4,f"Shift: {shift_str}   |   {hint}")
                c.font=fn(False,8,'2E7D32' if is_open else '1A6B3A')
                c.fill=fl('E8F0E8' if is_open else 'E8F4FD'); c.alignment=al('left'); c.border=tb()
                for rr2,lbl,bg in [(RS,'Status','F8FBFF'),(RI,'InTime','FFFFF0'),(RO,'OutTime','F0FFF0'),
                                   (RD,'Duration','F0F0FF'),(RL,'Late By','FFF3E0'),(RE,'Early By','FFF8F0'),
                                   (ROT,'OT','F0FFF4'),(RDP,'Dur+OT','EEE8F7')]:
                    c=ws.cell(rr2,1,lbl); c.font=fn(True,8); c.fill=fl('EEF4FA'); c.alignment=al(wrap=False); c.border=tb()
                    ws.cell(rr2,2).fill=fl(bg); ws.cell(rr2,2).border=tb()
                    ws.cell(rr2,3).fill=fl(bg); ws.cell(rr2,3).border=tb()
                for col in range(1,LAST_COL+1): ws.cell(RSP,col).fill=fl('DDE8F0')

                si_ref=f"$B${HR}"; so_ref=f"$C${HR}"
                for d in range(1,29):
                    col=DAY1_COL+d-1; cl=get_column_letter(col); sun=d in SUNDAYS
                    dd=emp_att.get(d,{}); raw_in=dd.get('in'); raw_out=dd.get('out')
                    has_in=raw_in is not None; has_out=raw_out is not None
                    miss=(has_in and not has_out) or (not has_in and has_out)
                    in_bg='FFE0B2' if miss else ('F0FFF8' if is_open else 'FFFFF0')
                    out_bg='FFE0B2' if miss else ('F0FFF8' if is_open else 'F0FFF0')
                    c=ws.cell(RI,col)
                    if has_in: c.value=time_val(raw_in); c.number_format='h:mm'
                    c.fill=fl(in_bg); c.font=fn(size=8); c.alignment=al(wrap=False); c.border=tb()
                    c=ws.cell(RO,col)
                    if has_out: c.value=time_val(raw_out); c.number_format='h:mm'
                    c.fill=fl(out_bg); c.font=fn(size=8); c.alignment=al(wrap=False); c.border=tb()
                    ir=f"{cl}{RI}"; or_=f"{cl}{RO}"
                    gp_mins=get_gp_deduction(gate_passes,emp_code,d); gp_frac=gp_mins/1440.0
                    if is_open: raw_dur=(f'MAX(0,MIN({or_},TIME(16,0,0))-{ir})' if sun else f'MAX(0,{or_}-{ir})')
                    else:
                        eff=f"MAX({si_ref},IF({ir}<={si_ref}+{GRACE_F},{si_ref},{ir}))"
                        raw_dur=(f'MAX(0,MIN({or_},TIME(16,0,0))-({eff}))' if sun else f'MAX(0,MIN({or_},{so_ref})-({eff}))')
                    if gp_mins>0: df_=f'=IF(OR({ir}="",{or_}=""),"",MAX(0,{raw_dur}-{gp_frac}))'; dur_fill='FFD580'
                    else: df_=f'=IF(OR({ir}="",{or_}=""),"",{raw_dur})'; dur_fill='E8F5E9' if is_open else 'F0F0FF'
                    c=ws.cell(RD,col); c.value=df_; c.number_format='[h]:mm'
                    c.font=fn(size=8); c.fill=fl(dur_fill); c.alignment=al(wrap=False); c.border=tb()
                    lf=('=""' if (is_open or sun) else f'=IF({ir}="","",IF({ir}>{si_ref}+{GRACE_F},{ir}-{si_ref},""))')
                    c=ws.cell(RL,col); c.value=lf; c.number_format='[h]:mm'
                    c.font=fn(size=8,color='B85C00'); c.fill=fl('FFF3E0'); c.alignment=al(wrap=False); c.border=tb()
                    ef=('=""' if (is_open or sun) else f'=IF({or_}="","",IF({or_}<{so_ref},{so_ref}-{or_},""))')
                    c=ws.cell(RE,col); c.value=ef; c.number_format='[h]:mm'
                    c.font=fn(size=8,color='795548'); c.fill=fl('FFF8F0'); c.alignment=al(wrap=False); c.border=tb()
                    dr_ref=f"{cl}{RD}"
                    if sun: of='=""'
                    elif is_open: of=f'=IF({or_}="","",IF({dr_ref}>{so_ref}+{OT_F},{dr_ref}-{so_ref},""))'
                    else: of=f'=IF({or_}="","",IF({or_}-{so_ref}>{OT_F},{or_}-{so_ref},""))'
                    c=ws.cell(ROT,col); c.value=of; c.number_format='[h]:mm'
                    c.font=fn(size=8,color='2E7D32'); c.fill=fl('F0FFF4'); c.alignment=al(wrap=False); c.border=tb()
                    ot_ref=f"{cl}{ROT}"
                    dpf=(f'=IF(AND({dr_ref}="",{ot_ref}=""),"",IFERROR('
                         f'IF(ISNUMBER({dr_ref}),{dr_ref},0)+IF(ISNUMBER({ot_ref}),{ot_ref},0),""))')
                    c=ws.cell(RDP,col); c.value=dpf; c.number_format='[h]:mm'
                    c.font=fn(size=8,color='4A235A'); c.fill=fl('EEE8F7'); c.alignment=al(wrap=False); c.border=tb()
                    sf=(f'=IF(AND({ir}="",{or_}=""),"WO",IF(OR({ir}="",{or_}=""),"MISS","WOP"))' if sun
                        else f'=IF(AND({ir}="",{or_}=""),"A",IF(OR({ir}="",{or_}=""),"MISS","P"))')
                    c=ws.cell(RS,col); c.value=sf; c.font=fn(size=8)
                    c.fill=fl('F8FBFF'); c.alignment=al(wrap=False); c.border=tb()

                sc2=get_column_letter(DAY1_COL); ec2=get_column_letter(DAY_END)
                sst=f"{sc2}{RS}:{ec2}{RS}"; sdr=f"{sc2}{RD}:{ec2}{RD}"
                slr=f"{sc2}{RL}:{ec2}{RL}"; ser=f"{sc2}{RE}:{ec2}{RE}"; sor_r=f"{sc2}{ROT}:{ec2}{ROT}"
                def ss(rng): return f"SUMPRODUCT(IF(ISNUMBER({rng}),{rng},0))"
                sw=ss(sdr); so2=ss(sor_r); sl=ss(slr); se2=ss(ser)
                pc=f'COUNTIF({sst},"P")'; wc=f'COUNTIF({sst},"WOP")'
                fmls=[
                    f'=COUNTIF({sst},"P")',f'=COUNTIF({sst},"WOP")',f'=COUNTIF({sst},"WO")',
                    f'=COUNTIF({sst},"A")',f'=COUNTIF({sst},"MISS")',
                    f'=IF({sw}=0,"0:00",TEXT({sw},"[h]:mm"))',
                    f'=IF({so2}=0,"0:00",TEXT({so2},"[h]:mm"))',
                    f'=IF({sl}=0,"0:00",TEXT({sl},"[h]:mm"))',
                    f'=SUMPRODUCT(({slr}<>"")*ISNUMBER({slr})*1)',
                    f'=IF({se2}=0,"0:00",TEXT({se2},"[h]:mm"))',
                    f'=SUMPRODUCT(({ser}<>"")*ISNUMBER({ser})*1)',
                    f'=IF(({sw}+{so2})=0,"0:00",TEXT({sw}+{so2},"[h]:mm"))',
                    f'=IFERROR(IF({sw}=0,"0:00",TEXT({sw}/MAX({pc}+{wc},1),"[h]:mm")),"0:00")',
                    f'=IF(MAX({sw}+{so2}-{sl},0)=0,"0:00",TEXT(MAX({sw}+{so2}-{sl},0),"[h]:mm"))',
                    f'=IF(({sw}+{so2})=0,"0:00",TEXT({sw}+{so2},"[h]:mm"))',
                ]
                for si_idx,formula in enumerate(fmls):
                    col=SUM1_COL+si_idx
                    ws.merge_cells(start_row=RS,start_column=col,end_row=RDP,end_column=col)
                    c=ws.cell(RS,col,formula)
                    c.font=fn(True,9,'4A235A') if si_idx==14 else fn(True,9)
                    c.fill=fl('EEE8F7') if si_idx==14 else fl('EBF5FB')
                    c.alignment=al(); c.border=mb()
                current_row=RSP+1

        ws.column_dimensions['A'].width=10; ws.column_dimensions['B'].width=5.5; ws.column_dimensions['C'].width=5.5
        for d in range(1,29): ws.column_dimensions[get_column_letter(DAY1_COL+d-1)].width=6.2
        for si3,sw_ in enumerate([8,6,6,7,6,11,9,10,8,10,8,12,10,11,11]):
            ws.column_dimensions[get_column_letter(SUM1_COL+si3)].width=sw_
        ws.freeze_panes="D6"

    buf=io.BytesIO(); wb.save(buf); buf.seek(0); return buf

# ─────────────────────────────────────────────
# BUILD TIMING SHEET
# ─────────────────────────────────────────────
def build_timing_sheet(emp_df, attendance, company_filter=None, dept_filter=None):
    wb=Workbook(); wb.remove(wb.active)
    thin=Side(style='thin')
    def tb(): return Border(left=thin,right=thin,top=thin,bottom=thin)
    def fl(h): return PatternFill('solid',start_color=h)
    def fn(bold=False,size=9,color='000000'): return Font(name='Arial',bold=bold,size=size,color=color)
    def al(h='center',v='center',wrap=True): return Alignment(horizontal=h,vertical=v,wrap_text=wrap)

    filtered=emp_df.copy()
    if company_filter: filtered=filtered[filtered['Company']==company_filter]
    if dept_filter:    filtered=filtered[filtered['Department']==dept_filter]

    for company in filtered['Company'].unique():
        c_emps=filtered[filtered['Company']==company]
        ws=wb.create_sheet(title=str(company)[:31])
        ws.merge_cells('A1:I1')
        c=ws.cell(1,1,f"Timing Sheet — {company} — February 2026")
        c.font=fn(True,12,'FFFFFF'); c.fill=fl('1A3C6E'); c.alignment=al()
        ws.row_dimensions[1].height=22
        headers=['Code','Name','Department','Shift','Date','Day','In Time','Out Time','Duration']
        for ci,h in enumerate(headers,1):
            c=ws.cell(2,ci,h); c.font=fn(True,9,'FFFFFF')
            c.fill=fl('2D6099'); c.alignment=al(); c.border=tb()
        ws.row_dimensions[2].height=16
        row=3
        for _,emp in c_emps.iterrows():
            try: code=int(float(str(emp['EmpCode'])))
            except: code=str(emp['EmpCode'])
            name=str(emp['EmpName']); dept=str(emp['Department'])
            si_t=emp['ShiftInTime']; so_t=emp['ShiftOutTime']
            is_open=bool(emp.get('IsOpen',False)); std_hours=float(emp.get('StdHours',9.5))
            shift_lbl='Open' if is_open else (f"{si_t.strftime('%H:%M')}–{so_t.strftime('%H:%M')}" if si_t and so_t else '?')
            emp_att=attendance.get(code,{})
            for d in range(1,29):
                dd=emp_att.get(d,{}); in_t=dd.get('in'); out_t=dd.get('out')
                r=calc_day(d,in_t,out_t,si_t,so_t,is_open,std_hours)
                if d in SUNDAYS:       bg='FFF3E0'
                elif d in SATURDAYS:   bg='F1F8E9'
                elif r['status']=='A': bg='FFEBEE'
                elif r['status']=='MISS': bg='FFF8E1'
                else: bg='FFFFFF' if row%2==0 else 'F8FBFF'
                vals=[code,name,dept,shift_lbl,f"Feb {d:02d}",DAY_ABBR[d],
                      in_t.strftime('%H:%M') if in_t else ('—' if r['status']=='A' else ''),
                      out_t.strftime('%H:%M') if out_t else ('—' if r['status']=='A' else ''),
                      m2hm(r['work']) if r['work']>0 else r['status']]
                for ci,val in enumerate(vals,1):
                    c=ws.cell(row,ci,val); c.font=fn(size=8)
                    c.fill=fl(bg); c.alignment=al(wrap=False); c.border=tb()
                row+=1
        for ci,w in enumerate([8,22,14,10,8,5,8,8,8],1):
            ws.column_dimensions[get_column_letter(ci)].width=w
        ws.freeze_panes='A3'
    buf=io.BytesIO(); wb.save(buf); buf.seek(0); return buf

# ─────────────────────────────────────────────
# SESSION STATE INIT
# ─────────────────────────────────────────────
for key,default in [('emp_df',None),('attendance',None),('summary_df',None),
                    ('miss_edits',{}),('gate_passes',[]),
                    ('open_shift_overrides',set()),('issues',[])]:
    if key not in st.session_state or st.session_state[key] is None:
        st.session_state[key] = default

# ─────────────────────────────────────────────
# SIDEBAR
# ─────────────────────────────────────────────
with st.sidebar:
    st.markdown("## 📋 Attendance System")
    st.markdown("---")
    st.markdown("### Upload Files")
    master_file=st.file_uploader("👥 Employee Master (.xlsx)", type=['xlsx','xls'], key='master_up')
    att_file   =st.file_uploader("🕐 Attendance Data (.xls/.xlsx)", type=['xlsx','xls'], key='att_up')

    if master_file and att_file:
        if st.button("⚙️ Process Files", use_container_width=True, type='primary'):
            with st.spinner("Loading & validating..."):
                try:
                    emp_df     = load_master(master_file)
                    attendance = load_attendance(att_file)
                    for code in st.session_state.open_shift_overrides:
                        emp_df.loc[emp_df['EmpCode']==code,'IsOpen']=True
                    issues=run_data_checks(emp_df,attendance)
                    st.session_state.emp_df     = emp_df
                    st.session_state.attendance = attendance
                    st.session_state.summary_df = compute_summary(emp_df,attendance,st.session_state.gate_passes)
                    st.session_state.miss_edits = {}
                    st.session_state.issues     = issues
                    cc=sum(1 for i in issues if '🔴' in i['severity'])
                    wc=sum(1 for i in issues if '🟡' in i['severity'])
                    mc=sum(1 for i in issues if '🟠' in i['severity'])
                    st.success(f"✅ {len(emp_df)} employees loaded")
                    if cc: st.error(f"🔴 {cc} critical issues — check Issues tab!")
                    if wc: st.warning(f"🟡 {wc} warnings")
                    if mc: st.warning(f"🟠 {mc} MISS punches")
                except Exception as e:
                    st.error(f"Error: {e}"); st.exception(e)

    st.markdown("---")
    if st.session_state.emp_df is not None:
        df=st.session_state.emp_df; issues=st.session_state.issues or []
        st.markdown(f"**Employees:** {len(df)}")
        st.markdown(f"**Companies:** {df['Company'].nunique()}")
        st.markdown(f"**Departments:** {df['Department'].nunique()}")
        cc=sum(1 for i in issues if '🔴' in i['severity'])
        if cc: st.error(f"⚠️ {cc} critical issues need attention!")

# ─────────────────────────────────────────────
# MAIN HEADER
# ─────────────────────────────────────────────
st.markdown('<div class="main-header"><h1>📋 Monthly Attendance Report System</h1>'
            '<p>Feb 2026 — Formula-Enabled Excel Generator</p></div>', unsafe_allow_html=True)

if st.session_state.emp_df is None:
    st.info("👈 Upload files in the sidebar to get started.")
    with st.expander("📖 How to use"):
        st.markdown("""
        1. Upload **Employee Master** and **Attendance XLS** in the sidebar
        2. Click **Process Files**
        3. **⚠️ Check Issues tab first** — fix critical problems before using salary data
        4. Use **Open Shift Override** tab for employees with irregular hours
        5. Fix any MISS punches, add Gate Passes, then download reports
        """)
    st.stop()

emp_df     = st.session_state.emp_df
attendance = st.session_state.attendance
summary_df = st.session_state.summary_df
issues     = st.session_state.issues or []

# Apply MISS edits
att_working={k:{d:dict(v) for d,v in v2.items()} for k,v2 in attendance.items()}
for (code,day),edits in st.session_state.miss_edits.items():
    if code not in att_working: att_working[code]={}
    if day not in att_working[code]: att_working[code][day]={'in':None,'out':None}
    if 'in'  in edits: att_working[code][day]['in']  = edits['in']
    if 'out' in edits: att_working[code][day]['out'] = edits['out']

critical_cnt=sum(1 for i in issues if '🔴' in i['severity'])
issue_badge=f" ({len(issues)})" if issues else ""

tab1,tab2,tab3,tab4,tab5,tab6,tab7,tab8 = st.tabs([
    "📊 Dashboard",
    "👥 Employee Master",
    "👁️ View Report",
    "✏️ Fix MISS Punches",
    "🔓 Open Shift Override",
    "🚪 Gate Pass",
    f"⚠️ Issues{issue_badge}",
    "⬇️ Download"
])

# ═══════════════════════════════════════════════
# TAB 1 — DASHBOARD
# ═══════════════════════════════════════════════
with tab1:
    st.subheader("📊 Attendance Summary Dashboard")
    if critical_cnt:
        st.error(f"⚠️ **{critical_cnt} critical issues** — review the Issues tab before using salary data!")

    c1,c2,c3,c4,c5=st.columns(5)
    c1.metric("👥 Employees",     len(summary_df))
    c2.metric("✅ Present Days",  summary_df['Present'].sum())
    c3.metric("❌ Absent Days",   summary_df['Absent'].sum())
    c4.metric("🟠 MISS Punches",  summary_df['MISS'].sum())
    c5.metric("⏱️ Avg Work/Day", m2hm(int(summary_df['_work_mins'].mean())))

    st.markdown("---")
    col1,col2=st.columns(2)
    with col1:
        comp_data=summary_df.groupby('Company')[['Present','WOP','Absent','MISS']].sum().reset_index()
        fig=px.bar(comp_data.melt(id_vars='Company',var_name='Status',value_name='Days'),
                   x='Company',y='Days',color='Status',barmode='group',
                   color_discrete_map={'Present':'#1976D2','WOP':'#388E3C','Absent':'#D32F2F','MISS':'#F57C00'},
                   title="Attendance by Company")
        fig.update_layout(height=350,margin=dict(l=0,r=0,t=40,b=0))
        st.plotly_chart(fig,use_container_width=True)
    with col2:
        late_df=summary_df[['Name','Late Days']].nlargest(10,'Late Days')
        fig2=px.bar(late_df,x='Name',y='Late Days',title="Top 10 Late Employees",
                    color='Late Days',color_continuous_scale='Oranges')
        fig2.update_layout(height=350,margin=dict(l=0,r=0,t=40,b=0),xaxis_tickangle=-30)
        st.plotly_chart(fig2,use_container_width=True)

    col3,col4=st.columns(2)
    with col3:
        ot_data=summary_df.groupby('Company')['_ot_mins'].sum().reset_index()
        ot_data['OT Hrs']=ot_data['_ot_mins']/60
        fig3=px.pie(ot_data,names='Company',values='OT Hrs',title="OT Hours by Company")
        fig3.update_layout(height=320,margin=dict(l=0,r=0,t=40,b=0))
        st.plotly_chart(fig3,use_container_width=True)
    with col4:
        summary_df['Work Hrs Num']=summary_df['_work_mins']/60
        fig4=px.histogram(summary_df,x='Work Hrs Num',nbins=20,title="Work Hours Distribution",
                          color_discrete_sequence=['#1565C0'])
        fig4.update_layout(height=320,margin=dict(l=0,r=0,t=40,b=0))
        st.plotly_chart(fig4,use_container_width=True)

    st.markdown("---")
    st.subheader("📋 Full Summary Table")
    fc1,fc2,fc3=st.columns(3)
    sel_company=fc1.multiselect("Company",    options=summary_df['Company'].unique(),    default=list(summary_df['Company'].unique()))
    sel_dept   =fc2.multiselect("Department", options=summary_df['Department'].unique(), default=list(summary_df['Department'].unique()))
    show_miss  =fc3.checkbox("Show only MISS employees")
    filtered=summary_df[summary_df['Company'].isin(sel_company)&summary_df['Department'].isin(sel_dept)]
    if show_miss: filtered=filtered[filtered['MISS']>0]
    display_cols=['Code','Name','Company','Department','Type','Shift','Present','WOP','W/Off',
                  'Absent','MISS','Work Hrs','OT Hrs','Gate Pass Hrs','GP Days',
                  'Dur+OT','Late Hrs','Late Days','Net Work Hrs','Avg Hrs/Day']
    st.dataframe(filtered[display_cols],use_container_width=True,hide_index=True,
                 column_config={'Present':st.column_config.NumberColumn('✅ P',width='small'),
                                'Absent':st.column_config.NumberColumn('❌ A',width='small'),
                                'MISS':st.column_config.NumberColumn('🟠 M',width='small'),
                                'WOP':st.column_config.NumberColumn('🟢 WOP',width='small')})

# ═══════════════════════════════════════════════
# TAB 2 — EMPLOYEE MASTER
# ═══════════════════════════════════════════════
with tab2:
    st.subheader("👥 Employee Master Data")
    st.caption("Exactly what was read from your uploaded Employee Master Excel file.")
    if critical_cnt:
        st.error(f"⚠️ {critical_cnt} critical issues found — see Issues tab.")

    mc1,mc2,mc3=st.columns(3)
    m_comp=mc1.multiselect("Company",    options=emp_df['Company'].unique(),    default=list(emp_df['Company'].unique()),   key='m_comp')
    m_dept=mc2.multiselect("Department", options=emp_df['Department'].unique(), default=list(emp_df['Department'].unique()),key='m_dept')
    m_type=mc3.radio("Shift Type",['All','🔒 Fixed','🔓 Open'],horizontal=True,key='m_type')

    m_filtered=emp_df[emp_df['Company'].isin(m_comp)&emp_df['Department'].isin(m_dept)].copy()
    if m_type=='🔒 Fixed': m_filtered=m_filtered[~m_filtered['IsOpen']]
    if m_type=='🔓 Open':  m_filtered=m_filtered[m_filtered['IsOpen']]

    def fmt_shift(row):
        if row['IsOpen']: return f"🔓 Open Shift (Std: {row['StdHours']}h)"
        si=row['ShiftInTime']; so=row['ShiftOutTime']
        if si and so: return f"🔒 {si.strftime('%H:%M')} – {so.strftime('%H:%M')}"
        return "⚠️ No Shift Times"

    disp=m_filtered[['EmpCode','EmpName','Company','Department','Hour','Timing','Shift']].copy()
    disp['Shift Info']=m_filtered.apply(fmt_shift,axis=1)
    disp.rename(columns={'EmpCode':'Code','EmpName':'Name','Hour':'Std Hrs'},inplace=True)
    st.markdown(f"**{len(m_filtered)} employees shown**")
    st.dataframe(disp[['Code','Name','Company','Department','Std Hrs','Timing','Shift Info']],
                 use_container_width=True,hide_index=True,
                 column_config={'Code':st.column_config.NumberColumn('Code',width='small'),
                                'Std Hrs':st.column_config.NumberColumn('Std Hrs',width='small'),
                                'Shift Info':st.column_config.TextColumn('Shift',width='large')})

    st.markdown("---")
    s1,s2,s3,s4=st.columns(4)
    s1.metric("Total Employees", len(emp_df))
    s2.metric("Fixed Shift",     int((~emp_df['IsOpen']).sum()))
    s3.metric("Open Shift",      int(emp_df['IsOpen'].sum()))
    s4.metric("Departments",     emp_df['Department'].nunique())

    orphan_codes=set(attendance.keys())-set(emp_df['EmpCode'].tolist())
    if orphan_codes:
        st.markdown("---")
        st.error(f"### 🔴 {len(orphan_codes)} Employees in Attendance but NOT in Master")
        st.markdown("These employees **cannot be processed for salary**. Add them to Employee Master and re-upload.")
        st.dataframe(pd.DataFrame([{'Employee Code':c,'Action':'Add to Employee Master Excel'} for c in sorted(orphan_codes)]),
                     use_container_width=True,hide_index=True)

# ═══════════════════════════════════════════════
# TAB 3 — VIEW REPORT
# ═══════════════════════════════════════════════
with tab3:
    st.subheader("👁️ View Attendance Report")
    companies=list(emp_df['Company'].unique())
    sel_comp=st.selectbox("Company",companies,key='vr_comp')
    c_emps=emp_df[emp_df['Company']==sel_comp]
    sel_dept2=st.selectbox("Department",list(c_emps['Department'].unique()),key='vr_dept')
    d_emps=c_emps[c_emps['Department']==sel_dept2]

    for _,emp in d_emps.iterrows():
        try: code=int(float(str(emp['EmpCode'])))
        except: code=str(emp['EmpCode'])
        si_t=emp['ShiftInTime']; so_t=emp['ShiftOutTime']
        is_open=bool(emp.get('IsOpen',False)); std_hours=float(emp.get('StdHours',9.5))
        emp_att=att_working.get(code,{})
        shift_lbl=(f"Open Shift (Std:{std_hours}h)" if is_open
                   else f"{si_t.strftime('%H:%M') if si_t else '?'} – {so_t.strftime('%H:%M') if so_t else '?'}")
        with st.expander(f"{'🔓' if is_open else '🔒'} {code} : {emp['EmpName']}  |  {shift_lbl}"):
            rows_v={'Status':[],'InTime':[],'OutTime':[],'Duration':[],'OT':[],'Dur+OT':[]}
            for d in range(1,29):
                dd=emp_att.get(d,{}); r=calc_day(d,dd.get('in'),dd.get('out'),si_t,so_t,is_open,std_hours)
                in_t=dd.get('in'); out_t=dd.get('out')
                rows_v['Status'].append(r['status'])
                rows_v['InTime'].append(in_t.strftime('%H:%M') if in_t else '')
                rows_v['OutTime'].append(out_t.strftime('%H:%M') if out_t else '')
                rows_v['Duration'].append(m2hm(r['work']) if r['work']>0 else '')
                rows_v['OT'].append(m2hm(r['ot']) if r['ot']>0 else '')
                rows_v['Dur+OT'].append(m2hm(r['work']+r['ot']) if (r['work']+r['ot'])>0 else '')
            view_df=pd.DataFrame.from_dict(rows_v,orient='index',
                                           columns=[f"{d} {DAY_ABBR[d][:2]}" for d in range(1,29)])
            def color_status(val):
                cm={'P':'background-color:#E3F2FD;color:#0D47A1;font-weight:bold',
                    'A':'background-color:#FFCDD2;color:#B71C1C;font-weight:bold',
                    'WOP':'background-color:#C8E6C9;color:#1B5E20;font-weight:bold',
                    'WO':'background-color:#EEEEEE;color:#555;font-weight:bold',
                    'MISS':'background-color:#FFE0B2;color:#E65100;font-weight:bold'}
                return cm.get(val,'')
            st.dataframe(view_df.style.applymap(color_status,subset=pd.IndexSlice[['Status'],:]),
                         use_container_width=True)
            sc1,sc2,sc3,sc4,sc5,sc6,sc7=st.columns(7)
            sr=summary_df[summary_df['Code']==code]
            if not sr.empty:
                row=sr.iloc[0]
                sc1.metric("Present", row['Present']); sc2.metric("WOP",     row['WOP'])
                sc3.metric("Absent",  row['Absent']);  sc4.metric("MISS",    row['MISS'])
                sc5.metric("Work Hrs",row['Work Hrs']); sc6.metric("OT Hrs", row['OT Hrs'])
                sc7.metric("Dur+OT",  row['Dur+OT'])

# ═══════════════════════════════════════════════
# TAB 4 — FIX MISS
# ═══════════════════════════════════════════════
with tab4:
    st.subheader("✏️ Fix MISS Punches")
    st.info("🟠 Enter the missing In or Out time and click 💾 Save.")
    miss_list=[]
    for _,emp in emp_df.iterrows():
        try: code=int(float(str(emp['EmpCode'])))
        except: code=str(emp['EmpCode'])
        emp_att=att_working.get(code,{})
        for d in range(1,29):
            dd=emp_att.get(d,{}); in_t=dd.get('in'); out_t=dd.get('out')
            if (in_t is None)!=(out_t is None):
                miss_list.append({'Code':code,'Name':str(emp['EmpName']),
                                  'Company':str(emp['Company']),'Department':str(emp['Department']),
                                  'Day':d,'InTime':in_t.strftime('%H:%M') if in_t else '❌ MISSING',
                                  'OutTime':out_t.strftime('%H:%M') if out_t else '❌ MISSING',
                                  '_missing':'in' if in_t is None else 'out'})
    if not miss_list:
        st.success("✅ No MISS punches found!")
    else:
        st.warning(f"**{len(miss_list)} MISS punches** across {len(set(m['Code'] for m in miss_list))} employees")
        mdf=pd.DataFrame(miss_list)
        fc1,fc2=st.columns(2)
        f_comp=fc1.selectbox("Filter Company",['All']+list(mdf['Company'].unique()),key='miss_comp')
        f_dept=fc2.selectbox("Filter Dept",   ['All']+list(mdf['Department'].unique()),key='miss_dept')
        fm=mdf.copy()
        if f_comp!='All': fm=fm[fm['Company']==f_comp]
        if f_dept!='All': fm=fm[fm['Department']==f_dept]
        for _,mr in fm.iterrows():
            code=mr['Code']; day=mr['Day']; ms=mr['_missing']
            ca,cb,cc,cd,ce=st.columns([3,2,2,2,1])
            ca.markdown(f"**{code} : {mr['Name']}**  \n`Feb {day:02d} ({DAY_ABBR[day]})`")
            cb.markdown(f"**In:** `{mr['InTime']}`"); cc.markdown(f"**Out:** `{mr['OutTime']}`")
            nt=cd.text_input(f"Enter {'In' if ms=='in' else 'Out'}Time",placeholder="e.g. 09:05",
                             key=f"miss_{code}_{day}_{ms}",label_visibility='collapsed')
            if ce.button("💾",key=f"save_{code}_{day}"):
                if nt.strip():
                    t=parse_time_str(nt.strip())
                    if t:
                        ek=(code,day)
                        if ek not in st.session_state.miss_edits: st.session_state.miss_edits[ek]={}
                        st.session_state.miss_edits[ek][ms]=t
                        att_working[code][day][ms]=t
                        st.session_state.summary_df=compute_summary(emp_df,att_working,st.session_state.gate_passes)
                        st.success(f"✅ Saved {nt} for {mr['Name']} Day {day}"); st.rerun()
                    else: st.error("Invalid time. Use HH:MM")
        if st.session_state.miss_edits:
            st.markdown("---")
            st.success(f"✅ {len(st.session_state.miss_edits)} edits saved")
            if st.button("🔄 Reset All Edits"):
                st.session_state.miss_edits={}
                st.session_state.summary_df=compute_summary(emp_df,attendance,st.session_state.gate_passes)
                st.rerun()

# ═══════════════════════════════════════════════
# TAB 5 — OPEN SHIFT OVERRIDE
# ═══════════════════════════════════════════════
with tab5:
    st.subheader("🔓 Open Shift Override")
    st.markdown("""
    Use this for employees who work **irregular hours** (e.g. 2h morning + 2h evening + 2h night shift).

    **Open Shift calculates:**
    - ✅ Duration = **Actual Out − In** (no fixed shift window capping)
    - ✅ OT = Duration exceeding **Standard Hours** (from Hour column in master)
    - ✅ **No Late / Early** deductions (irrelevant for irregular schedules)

    > 💡 If already marked "Open Shift" in your master Excel, they are handled automatically.
    > Use this tab only for **Fixed** shift employees who need Open Shift treatment this month.
    """)

    open_emps  = emp_df[emp_df['IsOpen']==True][['EmpCode','EmpName','Company','Department','StdHours']]
    fixed_emps = emp_df[emp_df['IsOpen']==False][['EmpCode','EmpName','Company','Department','StdHours']]

    col_a,col_b=st.columns(2)
    with col_a:
        st.markdown(f"### 🔓 Currently Open Shift ({len(open_emps)})")
        if not open_emps.empty:
            st.dataframe(open_emps.rename(columns={'EmpCode':'Code','EmpName':'Name','StdHours':'Std Hrs'}),
                         use_container_width=True,hide_index=True)
        else:
            st.info("No Open Shift employees in master.")

    with col_b:
        st.markdown(f"### 🔒 Fixed Shift — Override to Open ({len(fixed_emps)})")
        overrides=st.session_state.open_shift_overrides
        if not fixed_emps.empty:
            for _,er in fixed_emps.iterrows():
                code=er['EmpCode']; name=str(er['EmpName']); dept=str(er['Department']); std=er['StdHours']
                is_ov=code in overrides
                c1,c2=st.columns([4,1])
                c1.markdown(f"**{code}** — {name} `{dept}` Std:{std}h")
                toggled=c2.checkbox("Open",value=is_ov,key=f"os_{code}")
                if toggled and code not in overrides:
                    overrides.add(code)
                    emp_df.loc[emp_df['EmpCode']==code,'IsOpen']=True
                    st.session_state.open_shift_overrides=overrides
                    st.session_state.summary_df=compute_summary(emp_df,att_working,st.session_state.gate_passes)
                    st.rerun()
                elif not toggled and code in overrides:
                    overrides.discard(code)
                    emp_df.loc[emp_df['EmpCode']==code,'IsOpen']=False
                    st.session_state.open_shift_overrides=overrides
                    st.session_state.summary_df=compute_summary(emp_df,att_working,st.session_state.gate_passes)
                    st.rerun()

    if st.session_state.open_shift_overrides:
        st.markdown("---")
        st.success(f"✅ **{len(st.session_state.open_shift_overrides)} override(s) active:** "
                   f"{sorted(st.session_state.open_shift_overrides)}")
        st.caption("⚠️ Overrides apply to this session only. To make permanent, update Employee Master Excel.")

# ═══════════════════════════════════════════════
# TAB 6 — GATE PASS
# ═══════════════════════════════════════════════
with tab6:
    st.subheader("🚪 Gate Pass Manager")
    st.markdown("Gate pass duration is **automatically deducted from work hours**.")
    gate_passes=st.session_state.gate_passes

    with st.container(border=True):
        st.markdown("### ➕ Add New Gate Pass")
        ga,gb=st.columns(2)
        emp_options={f"{int(float(str(r['EmpCode'])))} — {r['EmpName']} ({r['Company']})":int(float(str(r['EmpCode'])))
                     for _,r in emp_df.iterrows()}
        sel_lbl=ga.selectbox("👤 Employee",list(emp_options.keys()),key="gp_emp")
        sel_code=emp_options[sel_lbl]; sel_name=sel_lbl.split('—')[1].split('(')[0].strip()
        gp_day=gb.selectbox("📅 Date",list(range(1,29)),format_func=lambda d:f"Feb {d:02d} ({DAY_ABBR[d]})",key="gp_day")
        gc,gd,ge,gf=st.columns(4)
        gp_out=gc.text_input("🚶 Out",placeholder="12:30",key="gp_out")
        gp_in =gd.text_input("🔙 In", placeholder="14:15",key="gp_in")
        gp_rsn=ge.text_input("📝 Reason",placeholder="Bank work",key="gp_rsn")
        gp_apv=gf.text_input("✅ Approved By",placeholder="Manager",key="gp_apv")
        if st.button("➕ Add Gate Pass",type="primary",use_container_width=True):
            ot=parse_time_str(gp_out.strip()) if gp_out.strip() else None
            it=parse_time_str(gp_in.strip())  if gp_in.strip()  else None
            if not ot or not it: st.error("❌ Enter valid Out and In times (HH:MM)")
            elif t2m(it)<=t2m(ot): st.error("❌ In Time must be after Out Time")
            else:
                dur=t2m(it)-t2m(ot)
                st.session_state.gate_passes.append({'code':sel_code,'name':sel_name,'day':gp_day,
                    'date_str':f"Feb {gp_day:02d} ({DAY_ABBR[gp_day]})",'gp_out':ot.strftime('%H:%M'),
                    'gp_in':it.strftime('%H:%M'),'duration_mins':dur,'duration_str':m2hm(dur),
                    'reason':gp_rsn.strip() or '—','approved_by':gp_apv.strip() or '—'})
                st.session_state.summary_df=compute_summary(emp_df,att_working,st.session_state.gate_passes)
                st.success(f"✅ Added for **{sel_name}** — {m2hm(dur)} deducted"); st.rerun()

    st.markdown("---")
    if not gate_passes: st.info("📭 No gate passes recorded yet.")
    else:
        st.markdown(f"### 📋 Gate Passes ({len(gate_passes)} total)")
        gf1,gf2=st.columns(2)
        gp_raw=pd.DataFrame(gate_passes)
        fe=gf1.selectbox("Filter Employee",['All']+sorted(gp_raw['name'].unique().tolist()),key="gp_fe")
        fd=gf2.selectbox("Filter Date",['All']+[f"Feb {d:02d} ({DAY_ABBR[d]})" for d in range(1,29)],key="gp_fd")
        fgp=[g for g in gate_passes
             if (fe=='All' or g['name']==fe) and
                (fd=='All' or g['day']==int(fd.split()[1]))]
        st.dataframe(pd.DataFrame([{'Code':g['code'],'Employee':g['name'],'Date':g['date_str'],
                                     'Out':g['gp_out'],'In':g['gp_in'],'Duration':g['duration_str'],
                                     'Reason':g['reason'],'Approved':g['approved_by']} for g in fgp]),
                     use_container_width=True,hide_index=True)
        gps={}
        for gp in gate_passes:
            k=(gp['code'],gp['name'])
            if k not in gps: gps[k]={'count':0,'total_mins':0,'days':set()}
            gps[k]['count']+=1; gps[k]['total_mins']+=gp['duration_mins']; gps[k]['days'].add(gp['day'])
        st.markdown("### 📊 Summary")
        st.dataframe(pd.DataFrame([{'Code':c,'Employee':n,'Passes':v['count'],'Days':len(v['days']),'Total Out':m2hm(v['total_mins'])}
                                    for (c,n),v in sorted(gps.items())]),use_container_width=True,hide_index=True)
        st.markdown("---")
        d1,d2=st.columns([3,1])
        dopts={f"#{i}|{g['name']}|{g['date_str']}|{g['gp_out']}–{g['gp_in']}":i for i,g in enumerate(gate_passes)}
        if dopts:
            ds=d1.selectbox("Select to delete",list(dopts.keys()),key="gp_del")
            if d2.button("🗑️ Delete",key="gp_del_btn"):
                st.session_state.gate_passes.pop(dopts[ds])
                st.session_state.summary_df=compute_summary(emp_df,att_working,st.session_state.gate_passes)
                st.rerun()
        if st.button("🔄 Clear ALL Gate Passes"):
            st.session_state.gate_passes=[]; st.rerun()
        cb=io.StringIO()
        pd.DataFrame([{'Code':g['code'],'Employee':g['name'],'Date':g['date_str'],'Out':g['gp_out'],
                        'In':g['gp_in'],'Duration':g['duration_str'],'Reason':g['reason'],'Approved':g['approved_by']}
                       for g in gate_passes]).to_csv(cb,index=False)
        st.download_button("📥 Export Gate Passes CSV",data=cb.getvalue(),file_name="GatePasses_Feb2026.csv",mime="text/csv")

# ═══════════════════════════════════════════════
# TAB 7 — ISSUES
# ═══════════════════════════════════════════════
with tab7:
    st.subheader("⚠️ Data Quality Issues")
    st.markdown("All issues that could affect **salary accuracy**. Fix 🔴 CRITICAL issues before downloading.")

    if not issues:
        st.success("✅ No issues found! Data looks clean.")
    else:
        sev_counts={}
        for i in issues: sev_counts[i['severity']]=sev_counts.get(i['severity'],0)+1
        ic1,ic2,ic3,ic4=st.columns(4)
        ic1.metric("🔴 Critical", sev_counts.get('🔴 CRITICAL',0))
        ic2.metric("🟡 Warnings", sev_counts.get('🟡 WARNING',0))
        ic3.metric("🟠 MISS",     sev_counts.get('🟠 MISS',0))
        ic4.metric("🔵 Info",     sev_counts.get('🔵 INFO',0))
        st.markdown("---")

        sev_filter=st.multiselect("Severity",['🔴 CRITICAL','🟡 WARNING','🟠 MISS','🔵 INFO'],
                                  default=['🔴 CRITICAL','🟡 WARNING','🟠 MISS'],key='isf')
        cat_filter=st.multiselect("Category",list({i['category'] for i in issues}),
                                  default=list({i['category'] for i in issues}),key='icf')
        fi=[i for i in issues if i['severity'] in sev_filter and i['category'] in cat_filter]
        st.markdown(f"**{len(fi)} issues shown**")
        st.markdown("---")

        for issue in fi:
            sev=issue['severity']
            box='issue-critical' if '🔴' in sev else ('issue-info' if '🔵' in sev else 'issue-box')
            st.markdown(
                f'<div class="{box}"><b>{sev} [{issue["category"]}]</b> — '
                f'Emp: <code>{issue["emp_code"]}</code> {issue["emp_name"]}<br>'
                f'{issue["detail"]}</div>', unsafe_allow_html=True)

        st.markdown("---")
        ie=pd.DataFrame([{'Severity':i['severity'],'Category':i['category'],'Emp Code':i['emp_code'],
                           'Emp Name':i['emp_name'],'Detail':i['detail']} for i in issues])
        cb2=io.StringIO(); ie.to_csv(cb2,index=False)
        st.download_button("📥 Export Issues CSV",data=cb2.getvalue(),
                           file_name="DataIssues_Feb2026.csv",mime="text/csv")

# ═══════════════════════════════════════════════
# TAB 8 — DOWNLOAD
# ═══════════════════════════════════════════════
with tab8:
    st.subheader("⬇️ Download Reports")
    if critical_cnt:
        st.error(f"⚠️ **{critical_cnt} critical issues** — fix them in Issues tab before downloading salary data!")

    # ── Attendance Excel ───────────────────────
    with st.container(border=True):
        st.markdown("### 📊 Full Attendance Report Excel")
        st.markdown("""
        - ✅ One sheet per Company, grouped by Department
        - ✅ 9 rows per employee: Status, InTime, OutTime, Duration, Late, Early, OT, Dur+OT
        - ✅ All formulas live — edit any punch time and everything recalculates
        - ✅ MISS punches highlighted orange
        - ✅ 15 summary columns including Net Work Hrs and Dur+OT
        - ✅ Gate pass deductions applied
        - ✅ Open Shift overrides included
        """)
        if st.session_state.gate_passes:
            st.info(f"🚪 {len(st.session_state.gate_passes)} gate passes included")
        if st.button("🔨 Generate Attendance Excel",type="primary",use_container_width=True,key='gen_att'):
            with st.spinner("Building Excel..."):
                try:
                    buf=build_excel(emp_df,att_working,st.session_state.gate_passes)
                    st.download_button("📥 Download Attendance_Feb2026_Report.xlsx",data=buf,
                                       file_name="Attendance_Feb2026_Report.xlsx",
                                       mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                       use_container_width=True,type="primary",key='dl_att')
                    st.success("✅ Ready! Click above to download.")
                except Exception as e: st.error(f"Error: {e}"); st.exception(e)

    st.markdown("---")

    # ── Timing Sheet ───────────────────────────
    with st.container(border=True):
        st.markdown("### 🖨️ Printable Timing Sheet Excel")
        st.markdown("""
        - ✅ One row per employee per day (all 28 days)
        - ✅ Shows In Time, Out Time, Duration, Status per row
        - ✅ Sunday / Saturday / Absent rows colour-coded
        - ✅ Filter by Company and Department
        - ✅ Print directly from Excel (Ctrl+P)
        """)
        ts1,ts2=st.columns(2)
        ts_co=ts1.selectbox("Company",['All']+list(emp_df['Company'].unique()),key='ts_co')
        ts_de=ts2.selectbox("Department",['All']+list(emp_df['Department'].unique()),key='ts_de')
        if st.button("🖨️ Generate Timing Sheet",type="primary",use_container_width=True,key='gen_ts'):
            with st.spinner("Building timing sheet..."):
                try:
                    buf2=build_timing_sheet(emp_df,att_working,
                                            None if ts_co=='All' else ts_co,
                                            None if ts_de=='All' else ts_de)
                    st.download_button("📥 Download TimingSheet_Feb2026.xlsx",data=buf2,
                                       file_name="TimingSheet_Feb2026.xlsx",
                                       mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                       use_container_width=True,type="primary",key='dl_ts')
                    st.success("✅ Timing sheet ready!")
                except Exception as e: st.error(f"Error: {e}"); st.exception(e)

    st.markdown("---")

    # ── Issues Export ──────────────────────────
    if issues:
        with st.container(border=True):
            st.markdown("### ⚠️ Issues Report")
            ie2=pd.DataFrame([{'Severity':i['severity'],'Category':i['category'],
                                'Emp Code':i['emp_code'],'Emp Name':i['emp_name'],'Detail':i['detail']}
                               for i in issues])
            cb3=io.StringIO(); ie2.to_csv(cb3,index=False)
            st.download_button("📥 Download Issues CSV",data=cb3.getvalue(),
                               file_name="DataIssues_Feb2026.csv",mime="text/csv",use_container_width=True)
